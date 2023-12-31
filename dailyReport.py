import openpyxl
import os

def write_edited_sheet(previous_sheet_title, current_date, current_stock_input, current_stock_output, previous_sheet, current_note, current_starting_stock):
    current_sheet = current_workbook.create_sheet(previous_sheet_title)
    # top header
    current_sheet['A1'] = 'Tanggal'
    current_sheet['B1'] = 'Input'
    current_sheet['C1'] = 'Output'

    current_sheet['A2'] = current_date
    current_sheet['B2'] = current_stock_input
    current_sheet['C2'] = current_stock_output

    # bottom header
    current_sheet['A4'] = 'Nomor'
    current_sheet['B4'] = 'Stock Awal'
    current_sheet['C4'] = 'Stock Akhir'
    current_sheet['D4'] = 'Tanggal'
    current_sheet['E4'] = 'Notes'

    i = 5
    # iterate previous xlsx to populate previous datas
    while previous_sheet.cell(row = i, column = 1).value is not None:
        current_sheet['A{}'.format(i)] = previous_sheet.cell(row = i, column = 1).value
        current_sheet['B{}'.format(i)] = previous_sheet.cell(row = i, column = 2).value
        current_sheet['C{}'.format(i)] = previous_sheet.cell(row = i, column = 3).value
        current_sheet['D{}'.format(i)] = previous_sheet.cell(row = i, column = 4).value
        current_sheet['E{}'.format(i)] = previous_sheet.cell(row = i, column = 5).value
        i += 1
    
    # write current data
    current_row = i
    current_sheet['A{}'.format(current_row)] = previous_sheet.cell(row = i-1 , column = 1).value + 1
    current_sheet['B{}'.format(current_row)] = current_starting_stock
    current_sheet['C{}'.format(current_row)] = current_starting_stock + current_stock_input - current_stock_output
    current_sheet['D{}'.format(current_row)] = current_date
    current_sheet['E{}'.format(current_row)] = current_note

def write_unedited_sheet(current_workbook, previous_workbook, current_sheet_name):
    unedited_sheet_names = [name for name in previous_workbook.sheetnames if name != current_sheet_name]
    for unedited_sheet_name in unedited_sheet_names:
        previous_unedited_sheet = previous_workbook[unedited_sheet_name]
        current_unedited_sheet = current_workbook.create_sheet(unedited_sheet_name)
        for row in previous_unedited_sheet.iter_rows(values_only=True):
            current_unedited_sheet.append(row)

try:
    print('Tanggal:')
    current_date = int(input())

    current_file_name = 'report-daily-tanggal-{}.xlsx'.format(current_date)

    # init previous_workbook from existing files
    previous_workbook = {}
    files = [f for f in os.listdir('.') if os.path.isfile(f)]
    date_iterator = current_date
    while date_iterator >= 0:
        file_name = 'report-daily-tanggal-{}.xlsx'.format(date_iterator)
        if file_name in files:
            previous_workbook = openpyxl.load_workbook(file_name)
            break
        date_iterator -= 1

    if previous_workbook == {}:
        raise FileNotFoundError('Tidak ditemukan file sebelum tanggal {}'.format(current_date))
        
    print('Nama Sheet:')
    current_sheet_name = input()

    previous_sheet = previous_workbook[current_sheet_name]
    last_history_row = 5
    while previous_sheet.cell(row = last_history_row, column = 1).value is not None:
        last_history_row += 1
    current_starting_stock = int(previous_sheet.cell(row = last_history_row - 1, column = 3).value)

    current_workbook = openpyxl.Workbook()
    current_workbook.remove(current_workbook.active)

    print('Input:')
    current_stock_input = int(input())
    print('Output:')
    current_stock_output = int(input())
    print('Notes:')
    current_note = input()

    write_edited_sheet(previous_sheet.title, current_date, current_stock_input, current_stock_output, previous_sheet, current_note, current_starting_stock)
    write_unedited_sheet(current_workbook, previous_workbook, current_sheet_name)

    current_workbook.save(current_file_name)
    current_workbook.close()
except FileNotFoundError as e:
    print(str(e))
except PermissionError:
    print('Error: Harap tutup file \'{}\' terlebih dahulu'.format(current_file_name))
except KeyError:
    print('Error: Sheet dengan nama \'{}\' tidak ditemukan'.format(current_sheet_name))